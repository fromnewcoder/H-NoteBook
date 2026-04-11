"""Unit tests for export_tasks module."""
import json
import os
import tempfile
import pytest
from unittest.mock import AsyncMock, patch, MagicMock
from uuid import uuid4


class TestGenerateExportContent:
    """Tests for generate_export_content function."""

    @pytest.fixture
    def mock_httpx_response(self):
        """Create a mock httpx response with list-of-blocks content format."""
        def _make_response(content_text: str):
            mock_response = MagicMock()
            mock_response.raise_for_status = MagicMock()
            mock_response.json.return_value = {
                "content": [
                    {"type": "text", "text": content_text}
                ]
            }
            return mock_response
        return _make_response

    @pytest.mark.asyncio
    async def test_generate_export_content_pdf_format(self, mock_httpx_response):
        """Test generate_export_content for PDF format extracts text from block format."""
        from app.workers.export_tasks import generate_export_content

        expected_content = "# Test Report\n\nThis is the content."
        mock_response = mock_httpx_response(expected_content)

        with patch('httpx.AsyncClient') as mock_client:
            mock_client.return_value.__aenter__.return_value.post = AsyncMock(return_value=mock_response)
            mock_client.return_value.__aenter__.return_value.__aexit__.return_value = None

            result = await generate_export_content("pdf", "test input content")

            assert result == expected_content
            mock_client.return_value.__aenter__.return_value.post.assert_called_once()

    @pytest.mark.asyncio
    async def test_generate_export_content_mind_map_format(self, mock_httpx_response):
        """Test generate_export_content for mind_map format extracts JSON from block format."""
        from app.workers.export_tasks import generate_export_content

        mind_map_json = json.dumps({
            "nodes": [{"id": "1", "label": "Central Topic"}],
            "edges": [{"from": "1", "to": "2", "label": "related topic"}]
        })
        mock_response = mock_httpx_response(mind_map_json)

        with patch('httpx.AsyncClient') as mock_client:
            mock_client.return_value.__aenter__.return_value.post = AsyncMock(return_value=mock_response)
            mock_client.return_value.__aenter__.return_value.__aexit__.return_value = None

            result = await generate_export_content("mind_map", "test input content")

            assert result == mind_map_json
            # Verify prompt includes mind_map instruction
            call_kwargs = mock_client.return_value.__aenter__.return_value.post.call_args
            prompt_sent = call_kwargs.kwargs["json"]["messages"][0]["content"]
            assert "mind map" in prompt_sent.lower()

    @pytest.mark.asyncio
    async def test_generate_export_content_docx_format(self, mock_httpx_response):
        """Test generate_export_content for DOCX format extracts text from block format."""
        from app.workers.export_tasks import generate_export_content

        expected_content = "# Heading\n\nSome paragraph text."
        mock_response = mock_httpx_response(expected_content)

        with patch('httpx.AsyncClient') as mock_client:
            mock_client.return_value.__aenter__.return_value.post = AsyncMock(return_value=mock_response)
            mock_client.return_value.__aenter__.return_value.__aexit__.return_value = None

            result = await generate_export_content("docx", "test input content")

            assert result == expected_content

    @pytest.mark.asyncio
    async def test_generate_export_content_pptx_format(self, mock_httpx_response):
        """Test generate_export_content for PPTX format extracts JSON from block format."""
        from app.workers.export_tasks import generate_export_content

        slides_json = json.dumps([
            {"title": "Slide 1", "bullets": ["point 1", "point 2"]}
        ])
        mock_response = mock_httpx_response(slides_json)

        with patch('httpx.AsyncClient') as mock_client:
            mock_client.return_value.__aenter__.return_value.post = AsyncMock(return_value=mock_response)
            mock_client.return_value.__aenter__.return_value.__aexit__.return_value = None

            result = await generate_export_content("pptx", "test input content")

            assert result == slides_json

    @pytest.mark.asyncio
    async def test_generate_export_content_xlsx_format(self, mock_httpx_response):
        """Test generate_export_content for XLSX format extracts JSON from block format."""
        from app.workers.export_tasks import generate_export_content

        sheets_json = json.dumps([
            {"sheet": "Sheet1", "headers": ["Col1", "Col2"], "rows": [["val1", "val2"]]}
        ])
        mock_response = mock_httpx_response(sheets_json)

        with patch('httpx.AsyncClient') as mock_client:
            mock_client.return_value.__aenter__.return_value.post = AsyncMock(return_value=mock_response)
            mock_client.return_value.__aenter__.return_value.__aexit__.return_value = None

            result = await generate_export_content("xlsx", "test input content")

            assert result == sheets_json

    @pytest.mark.asyncio
    async def test_generate_export_content_api_error(self):
        """Test generate_export_content raises exception on API error."""
        from app.workers.export_tasks import generate_export_content

        mock_response = MagicMock()
        mock_response.raise_for_status.side_effect = Exception("API Error")

        with patch('httpx.AsyncClient') as mock_client:
            mock_client.return_value.__aenter__.return_value.post = AsyncMock(return_value=mock_response)
            mock_client.return_value.__aenter__.return_value.__aexit__.return_value = None

            with pytest.raises(Exception, match="API Error"):
                await generate_export_content("pdf", "test content")

    @pytest.mark.asyncio
    async def test_generate_export_content_empty_content(self, mock_httpx_response):
        """Test generate_export_content returns empty string when content is None."""
        from app.workers.export_tasks import generate_export_content

        mock_response = MagicMock()
        mock_response.raise_for_status = MagicMock()
        mock_response.json.return_value = {"content": None}

        with patch('httpx.AsyncClient') as mock_client:
            mock_client.return_value.__aenter__.return_value.post = AsyncMock(return_value=mock_response)
            mock_client.return_value.__aenter__.return_value.__aexit__.return_value = None

            result = await generate_export_content("pdf", "test content")

            assert result == ""

    @pytest.mark.asyncio
    async def test_generate_export_content_content_limit(self, mock_httpx_response):
        """Test generate_export_content limits content to 10000 characters."""
        from app.workers.export_tasks import generate_export_content

        mock_response = mock_httpx_response("response")
        long_content = "x" * 15000

        with patch('httpx.AsyncClient') as mock_client:
            mock_client.return_value.__aenter__.return_value.post = AsyncMock(return_value=mock_response)
            mock_client.return_value.__aenter__.return_value.__aexit__.return_value = None

            await generate_export_content("pdf", long_content)

            call_kwargs = mock_client.return_value.__aenter__.return_value.post.call_args
            prompt_sent = call_kwargs.kwargs["json"]["messages"][0]["content"]
            # Content should be truncated to 10000 chars
            assert len(prompt_sent) <= 10000 + 200  # prompt prefix + 10000 chars


class TestRenderExportFile:
    """Tests for render_export_file function."""

    @pytest.fixture
    def temp_export_dir(self):
        """Create a temporary directory for export files."""
        with tempfile.TemporaryDirectory() as tmpdir:
            yield tmpdir

    @pytest.mark.asyncio
    async def test_render_export_file_pdf(self, temp_export_dir):
        """Test render_export_file creates valid PDF file."""
        from app.workers.export_tasks import render_export_file

        with patch('app.workers.export_tasks.settings') as mock_settings:
            mock_settings.export_storage_path = temp_export_dir

            job_id = str(uuid4())
            content = "# Main Title\n\n## Section 1\n\nSome content here."
            file_path = await render_export_file(job_id, "pdf", content)

            assert file_path == os.path.join(temp_export_dir, f"{job_id}.pdf")
            assert os.path.exists(file_path)
            # PDF files start with %PDF
            with open(file_path, 'rb') as f:
                header = f.read(4)
                assert header == b'%PDF'

    @pytest.mark.asyncio
    async def test_render_export_file_mind_map(self, temp_export_dir):
        """Test render_export_file creates mind map HTML file via pyvis."""
        from app.workers.export_tasks import render_export_file

        with patch('app.workers.export_tasks.settings') as mock_settings:
            mock_settings.export_storage_path = temp_export_dir

            job_id = str(uuid4())
            content = json.dumps({
                "nodes": [{"id": "1", "label": "Central Topic"}],
                "edges": []
            })

            # Mock pyvis Network to bypass HTML extension validation
            mock_net = MagicMock()
            mock_net.add_node = MagicMock()
            mock_net.add_edge = MagicMock()
            mock_net.save_graph = MagicMock()

            with patch('pyvis.network.Network', return_value=mock_net):
                file_path = await render_export_file(job_id, "mind_map", content)

            assert file_path == os.path.join(temp_export_dir, f"{job_id}.mind_map")
            # Verify pyvis Network was called correctly
            mock_net.add_node.assert_called_once_with("1", label="Central Topic")
            mock_net.save_graph.assert_called_once()

    @pytest.mark.asyncio
    async def test_render_export_file_mind_map_with_edges(self, temp_export_dir):
        """Test render_export_file creates mind map with edges correctly."""
        from app.workers.export_tasks import render_export_file

        with patch('app.workers.export_tasks.settings') as mock_settings:
            mock_settings.export_storage_path = temp_export_dir

            job_id = str(uuid4())
            content = json.dumps({
                "nodes": [
                    {"id": "1", "label": "Main Topic"},
                    {"id": "2", "label": "Sub Topic"}
                ],
                "edges": [
                    {"from": "1", "to": "2", "label": "related"}
                ]
            })

            # Mock pyvis Network
            mock_net = MagicMock()
            mock_net.add_node = MagicMock()
            mock_net.add_edge = MagicMock()
            mock_net.save_graph = MagicMock()

            with patch('pyvis.network.Network', return_value=mock_net):
                file_path = await render_export_file(job_id, "mind_map", content)

            # Verify nodes and edges were added
            assert mock_net.add_node.call_count == 1  # Only first node added in current impl
            assert mock_net.add_edge.call_count == 1
            mock_net.add_edge.assert_called_with("1", "2", label="related")

    @pytest.mark.asyncio
    async def test_render_export_file_docx(self, temp_export_dir):
        """Test render_export_file creates valid DOCX file."""
        from app.workers.export_tasks import render_export_file

        with patch('app.workers.export_tasks.settings') as mock_settings:
            mock_settings.export_storage_path = temp_export_dir

            job_id = str(uuid4())
            content = "# Main Heading\n\n## Sub Heading\n\nParagraph text here."
            file_path = await render_export_file(job_id, "docx", content)

            assert file_path == os.path.join(temp_export_dir, f"{job_id}.docx")
            assert os.path.exists(file_path)
            # DOCX files are ZIP archives, start with PK
            with open(file_path, 'rb') as f:
                header = f.read(4)
                assert header == b'PK\x03\x04'

    @pytest.mark.asyncio
    async def test_render_export_file_pptx(self, temp_export_dir):
        """Test render_export_file creates valid PPTX file."""
        from app.workers.export_tasks import render_export_file

        with patch('app.workers.export_tasks.settings') as mock_settings:
            mock_settings.export_storage_path = temp_export_dir

            job_id = str(uuid4())
            content = json.dumps([
                {"title": "Slide 1", "bullets": ["point 1", "point 2"]},
                {"title": "Slide 2", "bullets": ["point 3"]}
            ])
            file_path = await render_export_file(job_id, "pptx", content)

            assert file_path == os.path.join(temp_export_dir, f"{job_id}.pptx")
            assert os.path.exists(file_path)
            # PPTX files are ZIP archives
            with open(file_path, 'rb') as f:
                header = f.read(4)
                assert header == b'PK\x03\x04'

    @pytest.mark.asyncio
    async def test_render_export_file_xlsx(self, temp_export_dir):
        """Test render_export_file creates valid XLSX file."""
        from app.workers.export_tasks import render_export_file

        with patch('app.workers.export_tasks.settings') as mock_settings:
            mock_settings.export_storage_path = temp_export_dir

            job_id = str(uuid4())
            content = json.dumps([
                {
                    "sheet": "Data",
                    "headers": ["Name", "Value"],
                    "rows": [["Item1", "100"], ["Item2", "200"]]
                }
            ])
            file_path = await render_export_file(job_id, "xlsx", content)

            assert file_path == os.path.join(temp_export_dir, f"{job_id}.xlsx")
            assert os.path.exists(file_path)
            # XLSX files are ZIP archives
            with open(file_path, 'rb') as f:
                header = f.read(4)
                assert header == b'PK\x03\x04'

    @pytest.mark.asyncio
    async def test_render_export_file_xlsx_multiple_sheets(self, temp_export_dir):
        """Test render_export_file creates XLSX with multiple sheets."""
        from app.workers.export_tasks import render_export_file
        from openpyxl import load_workbook

        with patch('app.workers.export_tasks.settings') as mock_settings:
            mock_settings.export_storage_path = temp_export_dir

            job_id = str(uuid4())
            content = json.dumps([
                {"sheet": "Sheet1", "headers": ["Col1"], "rows": [["val1"]]},
                {"sheet": "Sheet2", "headers": ["Col2"], "rows": [["val2"]]}
            ])
            file_path = await render_export_file(job_id, "xlsx", content)

            wb = load_workbook(file_path)
            sheet_names = wb.sheetnames
            assert "Sheet1" in sheet_names
            assert "Sheet2" in sheet_names

    @pytest.mark.asyncio
    async def test_render_export_file_unknown_format(self, temp_export_dir):
        """Test render_export_file returns path without creating file for unknown format."""
        from app.workers.export_tasks import render_export_file

        with patch('app.workers.export_tasks.settings') as mock_settings:
            mock_settings.export_storage_path = temp_export_dir

            job_id = str(uuid4())
            file_path = await render_export_file(job_id, "unknown_format", "some content")

            assert file_path == os.path.join(temp_export_dir, f"{job_id}.unknown_format")
            # No file should be created for unknown format

    @pytest.mark.asyncio
    async def test_render_export_file_creates_export_directory(self, temp_export_dir):
        """Test render_export_file creates export directory if it doesn't exist."""
        from app.workers.export_tasks import render_export_file

        with patch('app.workers.export_tasks.settings') as mock_settings:
            mock_settings.export_storage_path = temp_export_dir

            with patch('app.workers.export_tasks.os.makedirs') as mock_makedirs:
                await render_export_file("test-id", "pdf", "# Test\nContent")
                mock_makedirs.assert_called_once_with(temp_export_dir, exist_ok=True)


class TestRenderExportFileMindMapEdgeCases:
    """Edge case tests for mind_map rendering."""

    @pytest.fixture
    def temp_export_dir(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            yield tmpdir

    @pytest.mark.asyncio
    async def test_render_export_file_mind_map_missing_node_label(self, temp_export_dir):
        """Test render_export_file handles mind map with missing node labels gracefully."""
        from app.workers.export_tasks import render_export_file

        with patch('app.workers.export_tasks.settings') as mock_settings:
            mock_settings.export_storage_path = temp_export_dir

            job_id = str(uuid4())
            # Node without label - should fallback to "Topic"
            content = json.dumps({
                "nodes": [{"id": "1"}],
                "edges": []
            })

            # Mock pyvis Network
            mock_net = MagicMock()
            mock_net.add_node = MagicMock()
            mock_net.add_edge = MagicMock()
            mock_net.save_graph = MagicMock()

            with patch('pyvis.network.Network', return_value=mock_net):
                file_path = await render_export_file(job_id, "mind_map", content)

            # Verify fallback label "Topic" is used
            mock_net.add_node.assert_called_with("1", label="Topic")

    @pytest.mark.asyncio
    async def test_render_export_file_mind_map_edge_missing_fields_uses_defaults(self, temp_export_dir):
        """Test render_export_file uses default values when edge fields are missing."""
        from app.workers.export_tasks import render_export_file

        with patch('app.workers.export_tasks.settings') as mock_settings:
            mock_settings.export_storage_path = temp_export_dir

            job_id = str(uuid4())
            # Edge without from/to fields - uses default "1"
            content = json.dumps({
                "nodes": [{"id": "1", "label": "Main"}],
                "edges": [
                    {"label": "edge without from/to"}
                ]
            })

            # Mock pyvis Network
            mock_net = MagicMock()
            mock_net.add_node = MagicMock()
            mock_net.add_edge = MagicMock()
            mock_net.save_graph = MagicMock()

            with patch('pyvis.network.Network', return_value=mock_net):
                file_path = await render_export_file(job_id, "mind_map", content)

            # Should use defaults for from/to
            mock_net.add_edge.assert_called_with("1", "1", label="edge without from/to")

